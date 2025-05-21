import os
import json
import csv
import pandas as pd
import logging
from typing import Dict, List, Optional, Tuple, Any, BinaryIO
import shutil
from datetime import datetime
import uuid
from fastapi import UploadFile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import settings

logger = logging.getLogger(__name__)

class FileHandler:
    """Utility class for handling file operations"""
    
    @staticmethod
    async def save_upload_file(file: UploadFile, user_id: str, file_type: str) -> str:
        """
        Save an uploaded file to disk
        
        Args:
            file: The uploaded file
            user_id: User ID to organize files
            file_type: Type of file (e.g., 'gst2b', 'purchase')
            
        Returns:
            str: Path to the saved file
        """
        # Create directory if it doesn't exist
        upload_dir = os.path.join(settings.UPLOAD_DIR, user_id)
        os.makedirs(upload_dir, exist_ok=True)
        
        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{file_type}_{timestamp}_{uuid.uuid4().hex[:8]}{os.path.splitext(file.filename)[1]}"
        
        file_path = os.path.join(upload_dir, filename)
        
        # Save file
        try:
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            logger.info(f"File saved successfully: {file_path}")
            return file_path
        except Exception as e:
            logger.error(f"Error saving file: {str(e)}")
            raise IOError(f"Failed to save file: {str(e)}")
        finally:
            # Make sure to close the file
            await file.close()
    
    @staticmethod
    def validate_json_file(file_path: str) -> bool:
        """
        Validate if a file is a valid JSON file with expected GST 2B structure
        
        Args:
            file_path: Path to the JSON file
            
        Returns:
            bool: True if valid, False otherwise
        """
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)
            
            # Check for expected structure
            if 'data' not in data:
                return False
            
            if 'docdata' not in data['data']:
                return False
            
            if 'b2b' not in data['data']['docdata']:
                return False
            
            return True
        except Exception as e:
            logger.error(f"JSON validation error: {str(e)}")
            return False
    
    @staticmethod
    def validate_csv_file(file_path: str) -> bool:
        """
        Validate if a file is a valid CSV file with expected purchase register structure
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            bool: True if valid, False otherwise
        """
        try:
            df = pd.read_csv(file_path)
            
            # Check for required columns
            required_columns = [
                'GSTIN of Supplier', 
                'Invoice Number', 
                'Invoice date', 
                'Invoice Value',
                'Taxable Value'
            ]
            
            for col in required_columns:
                if col not in df.columns:
                    logger.warning(f"Missing required column: {col}")
                    return False
            
            return True
        except Exception as e:
            logger.error(f"CSV validation error: {str(e)}")
            return False
    
    @staticmethod
    def generate_excel_report(data: pd.DataFrame, summary: Dict[str, Any], output_path: str) -> str:
        """
        Generate an Excel report from reconciliation results
        
        Args:
            data: DataFrame with reconciliation results
            summary: Summary statistics
            output_path: Directory to save the report
            
        Returns:
            str: Path to the generated Excel file
        """
        os.makedirs(output_path, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"gst_reconciliation_report_{timestamp}.xlsx"
        file_path = os.path.join(output_path, filename)
        
        # Create a new workbook and add worksheets
        wb = openpyxl.Workbook()
        
        # Format the main reconciliation sheet
        ws_reconciliation = wb.active
        ws_reconciliation.title = "Reconciliation Results"
        
        # Add a summary sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Add headers to reconciliation sheet
        headers = [
            "GSTIN", "Invoice Number", "Invoice Date", "Source", 
            "Taxable Value", "CGST", "SGST", "IGST", "Total Tax",
            "Match Status", "Remarks"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_reconciliation.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data to reconciliation sheet
        for idx, row in data.iterrows():
            row_num = idx + 2  # +2 because we start at row 2 (headers are row 1)
            
            ws_reconciliation.cell(row=row_num, column=1).value = row.get('gstin', '')
            ws_reconciliation.cell(row=row_num, column=2).value = row.get('invoice_number', '')
            ws_reconciliation.cell(row=row_num, column=3).value = row.get('invoice_date', '')
            ws_reconciliation.cell(row=row_num, column=4).value = row.get('source', '')
            ws_reconciliation.cell(row=row_num, column=5).value = row.get('taxable_value', 0)
            ws_reconciliation.cell(row=row_num, column=6).value = row.get('cgst', 0)
            ws_reconciliation.cell(row=row_num, column=7).value = row.get('sgst', 0)
            ws_reconciliation.cell(row=row_num, column=8).value = row.get('igst', 0)
            ws_reconciliation.cell(row=row_num, column=9).value = row.get('total_tax', 0)
            
            status = row.get('match_status', '')
            ws_reconciliation.cell(row=row_num, column=10).value = status
            
            # Add remarks based on match status
            remarks = ""
            if status == "match":
                remarks = "Eligible for ITC"
            elif status == "unmatch":
                remarks = "Mismatch - verify with vendor"
            elif status == "not in json":
                remarks = "Not found in GSTR-2B - ITC not eligible"
            elif status == "not in csv":
                remarks = "Not in purchase register - unclaimed ITC"
            
            ws_reconciliation.cell(row=row_num, column=11).value = remarks
            
            # Format cells based on match status
            status_cell = ws_reconciliation.cell(row=row_num, column=10)
            if status == "match":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status == "unmatch":
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif status == "not in json" or status == "not in csv":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Format the summary sheet
        summary_headers = ["Metric", "Value"]
        for col_num, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Add summary data
        summary_rows = [
            ("Total Invoices", summary.get('total_invoices', 0)),
            ("Matched Invoices", summary.get('matched_invoices', 0)),
            ("Unmatched Invoices", summary.get('unmatched_invoices', 0)),
            ("Missing in GSTR-2B", summary.get('missing_in_gst2b', 0)),
            ("Missing in Purchase Register", summary.get('missing_in_purchase', 0)),
            ("", ""),
            ("Total Tax in GSTR-2B", summary.get('total_tax_in_gst2b', 0)),
            ("Total Tax in Purchase Register", summary.get('total_tax_in_purchase', 0)),
            ("", ""),
            ("Eligible ITC", summary.get('eligible_itc', 0)),
            ("Excess Tax Claimed", summary.get('excess_tax_claimed', 0)),
            ("Unclaimed Tax", summary.get('unclaimed_tax', 0))
        ]
        
        for row_num, (metric, value) in enumerate(summary_rows, 2):
            ws_summary.cell(row=row_num, column=1).value = metric
            ws_summary.cell(row=row_num, column=2).value = value
            
            # Format currency cells
            if isinstance(value, (int, float)) and "Tax" in metric or "ITC" in metric:
                ws_summary.cell(row=row_num, column=2).number_format = "#,##0.00"
        
        # Auto-adjust column widths
        for worksheet in [ws_reconciliation, ws_summary]:
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(file_path)
        logger.info(f"Excel report generated: {file_path}")
        
        return file_path
    
    @staticmethod
    def export_to_csv(data: pd.DataFrame, output_path: str, filename_prefix: str = "reconciliation") -> str:
        """
        Export reconciliation results to CSV
        
        Args:
            data: DataFrame with reconciliation results
            output_path: Directory to save the CSV file
            filename_prefix: Prefix for the filename
            
        Returns:
            str: Path to the generated CSV file
        """
        os.makedirs(output_path, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.csv"
        file_path = os.path.join(output_path, filename)
        
        data.to_csv(file_path, index=False)
        logger.info(f"CSV report generated: {file_path}")
        
        return file_path
    
    @staticmethod
    def clean_up_old_files(directory: str, max_age_days: int = 7) -> int:
        """
        Delete files older than the specified age
        
        Args:
            directory: Directory to clean up
            max_age_days: Maximum age of files in days
            
        Returns:
            int: Number of files deleted
        """
        if not os.path.exists(directory):
            return 0
        
        files_deleted = 0
        current_time = datetime.now()
        
        for filename in os.listdir(directory):
            file_path = os.path.join(directory, filename)
            
            # Skip directories
            if os.path.isdir(file_path):
                continue
            
            # Get file creation/modification time
            file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
            age_days = (current_time - file_time).days
            
            if age_days > max_age_days:
                try:
                    os.remove(file_path)
                    files_deleted += 1
                    logger.info(f"Deleted old file: {file_path}")
                except Exception as e:
                    logger.error(f"Failed to delete file {file_path}: {str(e)}")
        
        return files_deleted